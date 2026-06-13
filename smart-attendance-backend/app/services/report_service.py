import io
from datetime import datetime
from typing import List, Dict, Any
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.course import Course
from app.models.student import Student
from app.schemas.reports import DefaulterResponse

class ReportService:
    @staticmethod
    def _create_pdf_doc(buffer, orientation="portrait"):
        pagesize = landscape(letter) if orientation == "landscape" else letter
        return SimpleDocTemplate(buffer, pagesize=pagesize, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)

    @staticmethod
    def generate_institution_attendance_pdf(institution_name: str, data: Dict, academic_year: str, semester: str) -> bytes:
        buffer = io.BytesIO()
        doc = ReportService._create_pdf_doc(buffer)
        styles = getSampleStyleSheet()
        elements = []

        # Header
        elements.append(Paragraph(f"<b>{institution_name}</b>", styles['Title']))
        elements.append(Paragraph(f"Institution Attendance Report", styles['Heading2']))
        elements.append(Paragraph(f"Academic Year: {academic_year} | Semester: {semester}", styles['Normal']))
        elements.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary Stats Table
        summary_data = [
            ["Metric", "Value"],
            ["Total Students", str(data.get("total_students", 0))],
            ["Total Sessions", str(data.get("total_sessions", 0))],
            ["Average Attendance", f"{data.get('avg_attendance', 0):.1f}%"]
        ]
        t = Table(summary_data, colWidths=[200, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # Department Breakdown
        elements.append(Paragraph("Department Breakdown", styles['Heading3']))
        dept_data = [["Department", "Code", "Avg Attendance %"]]
        for dept in data.get("departments", []):
            dept_data.append([dept["name"], dept["code"], f"{dept['avg_pct']:.1f}%"])
            
        t2 = Table(dept_data, colWidths=[200, 100, 150])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.steelblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t2)

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def generate_department_attendance_pdf(department, data: Dict, academic_year: str, semester: str) -> bytes:
        buffer = io.BytesIO()
        doc = ReportService._create_pdf_doc(buffer)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"<b>Department: {department.name} ({department.code})</b>", styles['Title']))
        elements.append(Paragraph(f"Department Attendance Report", styles['Heading2']))
        elements.append(Paragraph(f"Academic Year: {academic_year} | Semester: {semester}", styles['Normal']))
        elements.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary
        summary_data = [
            ["Metric", "Value"],
            ["Average Attendance", f"{data.get('avg_attendance', 0):.1f}%"]
        ]
        t = Table(summary_data, colWidths=[200, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        for prog in data.get("programmes", []):
            elements.append(Paragraph(f"Programme: {prog['name']} ({prog['code']})", styles['Heading3']))
            c_data = [["Course Code", "Course Title", "Avg Attendance %"]]
            for c in prog.get("courses", []):
                c_data.append([c["code"], c["title"], f"{c['avg_pct']:.1f}%"])
            
            if len(c_data) > 1:
                t2 = Table(c_data, colWidths=[100, 250, 100])
                t2.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.steelblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(t2)
            else:
                elements.append(Paragraph("No courses found.", styles['Normal']))
            elements.append(Spacer(1, 15))

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def generate_course_attendance_pdf(course: Course, sessions: List, students: List, attendance_records: List) -> bytes:
        buffer = io.BytesIO()
        doc = ReportService._create_pdf_doc(buffer, "landscape")
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"Course Attendance: {course.code} - {course.title}", styles['Title']))
        elements.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Data rows
        table_data = [["Student Name", "Student ID", "Present", "Total", "%"]]
        
        # Calculate stats for each student
        for student in students:
            student_records = [r for r in attendance_records if r.student_id == student.id]
            present_count = sum(1 for r in student_records if r.status.value == "present")
            total = len(sessions)
            pct = (present_count / total * 100) if total > 0 else 0
            
            table_data.append([
                student.name,
                student.student_id,
                str(present_count),
                str(total),
                f"{pct:.1f}%"
            ])

        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def generate_per_student_pdf(student: Student, courses_data: List) -> bytes:
        buffer = io.BytesIO()
        doc = ReportService._create_pdf_doc(buffer)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"Student Attendance Report", styles['Title']))
        elements.append(Paragraph(f"Name: {student.name} | ID: {student.student_id}", styles['Heading3']))
        elements.append(Spacer(1, 20))

        table_data = [["Course Code", "Course Title", "Present", "Total", "%", "Status"]]
        for c in courses_data:
            table_data.append([
                c["course_code"], c["course_title"], 
                str(c["sessions_present"]), str(c["sessions_total"]),
                f"{c['attendance_pct']:.1f}%", c["status"]
            ])

        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def generate_defaulters_pdf(defaulters: List[DefaulterResponse], institution_name: str, filters: Dict) -> bytes:
        buffer = io.BytesIO()
        doc = ReportService._create_pdf_doc(buffer)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"<b>{institution_name}</b>", styles['Title']))
        elements.append(Paragraph("Attendance Defaulters Report", styles['Heading2']))
        elements.append(Spacer(1, 20))

        table_data = [["Student Name", "ID", "Course Code", "Current %", "Required %", "Shortfall %"]]
        for d in defaulters:
            table_data.append([
                d.student_name, d.student_number, d.course_code,
                f"{d.current_pct:.1f}", str(d.threshold_pct), f"{d.shortfall:.1f}"
            ])

        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def generate_lecturer_activity_pdf(lecturers_data: List, institution_name: str) -> bytes:
        buffer = io.BytesIO()
        doc = ReportService._create_pdf_doc(buffer)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Lecturer Activity Report", styles['Title']))
        elements.append(Spacer(1, 20))

        table_data = [["Lecturer", "Staff ID", "Courses", "Sessions Conducted", "Avg Attendance %"]]
        for l in lecturers_data:
            table_data.append([
                l["name"], l["staff_id"], str(l["course_count"]), 
                str(l["sessions_conducted"]), f"{l['avg_attendance']:.1f}%"
            ])

        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def generate_institution_attendance_excel(data: Dict) -> bytes:
        wb = Workbook()
        
        # Sheet 1
        ws1 = wb.active
        ws1.title = "Summary Stats"
        ws1.append(["Metric", "Value"])
        ws1.append(["Total Students", data.get("total_students", 0)])
        ws1.append(["Total Sessions", data.get("total_sessions", 0)])
        ws1.append(["Average Attendance %", round(data.get("avg_attendance", 0), 2)])
        for cell in ws1["1:1"]: cell.font = Font(bold=True)

        # Sheet 2
        ws2 = wb.create_sheet("Department Breakdown")
        ws2.append(["Department", "Code", "Avg Attendance %"])
        for d in data.get("departments", []):
            ws2.append([d["name"], d["code"], round(d["avg_pct"], 2)])
        for cell in ws2["1:1"]: cell.font = Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def generate_course_attendance_excel(course: Course, sessions: List, students: List, attendance_records: List) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Course Attendance Matrix"

        # Headers
        headers = ["Student Name", "Student ID", "Total %"]
        for s in sessions:
            headers.append(f"{s.session_date.strftime('%Y-%m-%d')} ({s.label or s.session_code})")
        ws.append(headers)
        for cell in ws["1:1"]: cell.font = Font(bold=True)

        # Build records map for quick lookup: (student_id, session_id) -> status
        records_map = {(r.student_id, r.session_id): r.status.value for r in attendance_records}

        for student in students:
            row = [student.name, student.student_id]
            
            # Calculate total
            student_records = [r for r in attendance_records if r.student_id == student.id]
            present_count = sum(1 for r in student_records if r.status.value == "present")
            total = len(sessions)
            pct = (present_count / total * 100) if total > 0 else 0
            row.append(round(pct, 2))

            # Sessions
            for session in sessions:
                status = records_map.get((student.id, session.id), "absent")
                val = "P" if status == "present" else "A"
                row.append(val)
                
            ws.append(row)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def generate_defaulters_excel(defaulters: List[DefaulterResponse]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Defaulters"

        ws.append(["Student Name", "ID", "Course Title", "Course Code", "Current %", "Required %", "Shortfall %"])
        for cell in ws["1:1"]: cell.font = Font(bold=True)

        for d in defaulters:
            ws.append([
                d.student_name, d.student_number, d.course_title, d.course_code,
                round(d.current_pct, 2), d.threshold_pct, round(d.shortfall, 2)
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def generate_lecturer_activity_excel(data: List) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lecturer Activity"

        ws.append(["Lecturer Name", "Staff ID", "Courses", "Sessions Conducted", "Avg Attendance %"])
        for cell in ws["1:1"]: cell.font = Font(bold=True)

        for l in data:
            ws.append([
                l["name"], l["staff_id"], l["course_count"], 
                l["sessions_conducted"], round(l["avg_attendance"], 2)
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
